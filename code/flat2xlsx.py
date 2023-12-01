import json
import logging.handlers
import csv

import typer
import pyodbc
# import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook


def to_csv(rows: list, header: list, output_file: str, options: dict = None):
    try:
        clean_rows = []
        for row in rows:       
            clean_rows.append([(x.strip() if isinstance(x, str) else x) for x in row])
        
        quoting_options = {
            'ALL': csv.QUOTE_ALL,
            'MINIMAL': csv.QUOTE_MINIMAL,
            'NONNUMERIC': csv.QUOTE_NONNUMERIC,
            'NONE': csv.QUOTE_NONE
        }

        quoting_option = csv.QUOTE_MINIMAL
        if options:
            quoting_option = quoting_options[options['quoting']]
        
        with open(output_file, 'w') as f:
            # csv.writer(f, quoting=csv.QUOTE_MINIMAL, lineterminator='\n').writerows([header])
            # csv.writer(f, quoting=csv.QUOTE_MINIMAL, lineterminator='\n').writerows(clean_rows)
            csv.writer(f, quoting=quoting_option, lineterminator='\n').writerows([header])
            csv.writer(f, quoting=quoting_option, lineterminator='\n').writerows(clean_rows)
    
    except Exception as sub_ex:
        log_dog.error(sub_ex)
        raise


def to_excel(rows: list, header: list, output_file: str):
    try:
        wb = Workbook()
        sheet = wb.active

        col_ctr = 1
        for column in header:
            sheet.cell(row=1, column=col_ctr).value = column
            col_ctr += 1

        row_ctr = 2
        for row in rows:
            col_ctr = 1
            for col in row:
                if type(col) == 'str':
                    sheet.cell(row=row_ctr, column=col_ctr).value = col.strip()
                else:
                    sheet.cell(row=row_ctr, column=col_ctr).value = col     
                col_ctr += 1
            row_ctr += 1

        wb.save(output_file)
        wb.close()

    except Exception as sub_ex:
        log_dog.error(sub_ex)
        raise


def get_connection(dsn: str, con: dict) -> pyodbc.Connection:
    try:
        if dsn:
            con_str = f"FILEDSN={dsn}" if (dsn.find(":\\")) != -1 else f"DSN={dsn}"
        else:
            if con["trusted"]:
                user_password = f"trusted_connect={con['trusted']}"
            else:
                user_password = f"UID={con['user']};PWD={con['password']}"
            con_str = f"DRIVER={con['driver']};SERVER={con['server']};DATABASE={con['db']};{user_password}"

        return pyodbc.connect(con_str)

    except Exception as sub_ex:
        log_dog.error(sub_ex)
        raise


def get_cursor(connection: pyodbc.Connection) -> pyodbc.Cursor:
    try:
        return connection.cursor()

    except Exception as sub_ex:
        log_dog.error(sub_ex)
        raise


def get_data(cursor: pyodbc.Cursor, sql: str) -> tuple:
    try:
        cursor.execute(sql)
        cursor_rows = cursor.fetchall()
        cursor_header = [column[0] for column in cursor.description]

        return cursor_rows, cursor_header

    except Exception as sub_ex:
        log_dog.error(sub_ex)
        raise


def get_sql_commands(sql_file: str) -> str:
    try:
        with open(sql_file, newline="\n", encoding="utf-8") as sql_file:
            return " ".join(line.strip() for line in sql_file)

    except Exception as sub_ex:
        log_dog.error(sub_ex)
        raise


def get_owner_lookup(db_rows: list) -> dict:
    owner_dict = dict()

    try:
        for db_row in db_rows:
            acctno, owner, address1, address2, csz, cr = db_row
            owner_address = (
                f"{owner[0:30]:<30}{address1[0:30]:<30}{address2:<30}{csz:<48}{cr:<4}"
            )
            owner_dict[acctno] = owner_address
        return owner_dict

    except Exception as sub_ex:
        log_dog.error(sub_ex)
        raise


def set_up_error_log(log_file: str) -> object:
    err_log = logging.getLogger(__name__)
    err_log.setLevel(logging.ERROR)

    # bundle_dir = path.abspath(path.dirname(__file__))
    # path_to_log = path.join(bundle_dir, log_file)

    h1 = logging.FileHandler(log_file)
    f = logging.Formatter(
        "%(levelname)s %(asctime)s %(funcName)s %(lineno)d %(message)s"
    )
    h1.setFormatter(f)
    h1.setLevel(logging.ERROR)
    err_log.addHandler(h1)

    return err_log


log_dog = set_up_error_log("flat2xlsx.log")
G_FTP_PATH = None
G_INCREMENT_SEED = 0
G_FILE_NAME_FORMAT = None
SQL_CONFIG = None
LOOKUP_LOG = None


def main(
    type: str = typer.Option(..., "--type", "-t", help="Export Type"),
    config: str = typer.Option(..., "--config", "-c", help="Config File Name and Path"),
):
    try:
        with open(config, encoding="utf-8") as config_file:
            settings = json.load(config_file)
            # dsn = settings["dsn"]
            input_file = settings[type]["files"]["input"]
            output_file = settings[type]["files"]["output"]
            sql_file = settings[type]["files"]["sql"]
            template_file = settings[type]["files"]["template"]
            # connection is a placeholder for the connection credentials
            # connection = {}


        summary = []

        wbt = load_workbook(template_file)
        wst = wbt['Bill Print Detail']
        max_row = wst.max_row

        bill_layout = {}

        # Loop through each row and populate the dictionary
        for row in range(1, max_row + 1):
            key = wst.cell(row=row, column=1).value
            value = wst.cell(row=row, column=4).value
            if key and value:  # Skip rows where either key or value is None
                bill_layout[key] = value

        wb = Workbook()
        ws = wb.active

        column_ctr = 1

        for key, value in bill_layout.items():
            ws.cell(row=1, column=column_ctr).value = key
            column_ctr += 1

        start_record = 1

        with open(input_file, "r") as file:
            row_ctr = 2
            
            for line_number, line in enumerate(file, 1):
                if line[0:1] == 'D':
                    if line_number > start_record:
                        start_pos = 0
                        stop_pos = 0
                        row_dict = {}
                        for count, (key, value) in enumerate(bill_layout.items(), 1):
                            stop_pos = value
                            row_dict[key]=line[start_pos:stop_pos]
                            col = line[start_pos:stop_pos]
                            if isinstance(col, str):
                                ws.cell(row=row_ctr, column=count).value = col.strip()
                            else:
                                ws.cell(row=row_ctr, column=count).value = col     
                            start_pos = stop_pos
                    row_ctr += 1

        wb.save(output_file)
        wb.close()

    except Exception as main_ex:
        log_dog.exception("Main Process Exceptions: %s", main_ex)

if __name__ == "__main__":
    typer.run(main)

